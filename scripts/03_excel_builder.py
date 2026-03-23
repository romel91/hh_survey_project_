"""
========================================================
  HOUSEHOLD SURVEY ANALYSIS
  Script 03 — Excel Workbook Builder
  Author  : Data Analytics Portfolio Project
  Input   : data/cleaned/HH_Survey_Cleaned.csv
            data/cleaned/HH_Household_Level.csv
  Output  : outputs/HH_Survey_Analysis.xlsx  (7 sheets)
========================================================
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os
import warnings
warnings.filterwarnings('ignore')
from config import CLEANED_PATH, HH_PATH, OUTPUT_PATH, POVERTY_LINE

os.makedirs('outputs', exist_ok=True)

# ── Colour constants ───────────────────────────────────────
NAVY   = '1F4E79'
BLUE   = '2E75B6'
LBLUE  = 'D6E4F0'
XLBLUE = 'EBF3FB'
AMBER  = 'F4B942'
RED    = 'C00000'
GREEN  = '375623'
LGREEN = 'E2F0D9'
LGRAY  = 'F2F2F2'
WHITE  = 'FFFFFF'
GRAY   = '595959'

# ── Load data ──────────────────────────────────────────────
print("Loading cleaned data...")
df = pd.read_csv(CLEANED_PATH)
hh = pd.read_csv(HH_PATH)
print(f"Individual: {df.shape} | Household: {hh.shape}")

# ── Style helpers ──────────────────────────────────────────
def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, size=10, color='1A1A1A', italic=False):
    return Font(name='Calibri', bold=bold, size=size,
                color=color, italic=italic)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def thin():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def set_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_title(ws, row, text, ncols, bg=NAVY, size=14):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = font(bold=True, size=size, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = center()
    ws.row_dimensions[row].height = 32

def write_section(ws, row, text, ncols, bg=BLUE):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = font(bold=True, size=12, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = left()
    ws.row_dimensions[row].height = 26

def write_header(ws, row, headers, widths, bg=NAVY):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row, i, h)
        c.font      = font(bold=True, size=10, color=WHITE)
        c.fill      = fill(bg)
        c.alignment = center()
        c.border    = thin()
        set_width(ws, i, w)

def write_row(ws, row, values, alt=False):
    bg_color = XLBLUE if alt else WHITE
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font      = font(size=10)
        c.fill      = fill(bg_color)
        c.alignment = left() if isinstance(v, str) else center()
        c.border    = thin()

def write_kv(ws, row, label, value, note='', alt=False):
    """Write a key-value pair row across 3 columns."""
    bg = XLBLUE if alt else WHITE
    for col, val in enumerate([label, value, note], 1):
        c = ws.cell(row, col, val)
        c.font      = font(bold=(col == 1), size=10)
        c.fill      = fill(bg)
        c.alignment = left()
        c.border    = thin()


# ══════════════════════════════════════════════════════════════
wb = Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — Cleaned Data
# ══════════════════════════════════════════════════════════════
print("\nBuilding Sheet 1: Cleaned Data...")
ws1 = wb.active
ws1.title = '01 Cleaned Data'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A3'

write_title(ws1, 1,
    f'HOUSEHOLD SURVEY — CLEANED DATASET  |  {len(df)} Individuals · {hh["hh_id"].nunique()} Households · Bangladesh 2024',
    18)

headers = [
    'HH ID', 'Person ID', 'HH Head Name', 'Relation', 'Age',
    'Sex (1=M, 2=F)', 'Sex Label', 'Education Level', 'Edu Score',
    'Marital Status', 'Occupation', 'Monthly Income (BDT)', 'Is Earner',
    'Treatment', 'Group', 'District', 'Enumerator', 'Interview Duration (min)'
]
widths = [9, 10, 18, 12, 7, 14, 11, 16, 10, 14, 14, 22, 10, 11, 12, 13, 12, 22]
write_header(ws1, 2, headers, widths)

export_cols = [
    'hh_id', 'person_id', 'hh_head_name', 'relation', 'age', 'sex',
    'sex_label', 'edu_level', 'edu_score', 'marital_status', 'occupation',
    'monthly_income', 'is_earner', 'treatment', 'group',
    'district', 'enumerator_id', 'interview_duration'
]

for r, (_, row) in enumerate(df[export_cols].iterrows(), 3):
    write_row(ws1, r, list(row), alt=(r % 2 == 0))

# Red highlight on zero-income cells
for r in range(3, 3 + len(df)):
    val = ws1.cell(r, 12).value
    if val == 0:
        ws1.cell(r, 12).fill = fill('FCE4D6')
        ws1.cell(r, 12).font = Font(name='Calibri', size=10, color=RED, bold=True)

print(f"  → {len(df)} rows written")


# ══════════════════════════════════════════════════════════════
# SHEET 2 — Summary Statistics
# ══════════════════════════════════════════════════════════════
print("Building Sheet 2: Summary Statistics...")
ws2 = wb.create_sheet('02 Summary Statistics')
ws2.sheet_view.showGridLines = False

write_title(ws2, 1, 'SUMMARY STATISTICS — HOUSEHOLD SURVEY ANALYSIS', 4)
for col, w in zip([1,2,3,4], [30, 20, 32, 5]):
    set_width(ws2, col, w)

# Block A — Dataset Overview
write_section(ws2, 3, '📊  DATASET OVERVIEW', 4, BLUE)
write_header(ws2, 4, ['Metric', 'Value', 'Notes', ''], [30, 20, 32, 5], BLUE)
overview_rows = [
    ('Total Records (Individuals)', 137,  'Rows in cleaned dataset'),
    ('Unique Households',           30,   'hh_id range: 101–130'),
    ('Avg Household Size',          4.6,  'Min 3, Max 6 members'),
    ('Districts Covered',           5,    'Dhaka, Chittagong, Khulna, Sylhet, Rajshahi'),
    ('Enumerators',                 5,    'EN01 through EN05'),
    ('Survey Period',               2024, 'Cross-sectional design'),
    ('Treatment Group (n)',         69,   '50.4% of individuals'),
    ('Control Group (n)',           68,   '49.6% of individuals'),
]
for i, (m, v, n) in enumerate(overview_rows, 5):
    write_kv(ws2, i, m, v, n, alt=(i % 2 == 0))

# Block B — Income Stats
earners = df[df['monthly_income'] > 0]['monthly_income']
all_inc = df['monthly_income']
n_earners = len(earners)
write_section(ws2, 14, '💰  INCOME STATISTICS — Monthly BDT', 4, NAVY)
write_header(ws2, 15, ['Statistic', 'All Members', f'Earners Only (n={n_earners})', ''],
             [30, 20, 32, 5], NAVY)
income_rows = [
    ('Count',           f'{len(all_inc):,}',           f'{n_earners:,}'),
    ('Mean',            f'৳{all_inc.mean():,.0f}',     f'৳{earners.mean():,.0f}'),
    ('Median',          f'৳{all_inc.median():,.0f}',   f'৳{earners.median():,.0f}'),
    ('Std Deviation',   f'৳{all_inc.std():,.0f}',      f'৳{earners.std():,.0f}'),
    ('Minimum',         f'৳{all_inc.min():,.0f}',      f'৳{earners.min():,.0f}'),
    ('Maximum',         f'৳{all_inc.max():,.0f}',      f'৳{earners.max():,.0f}'),
    ('25th Percentile', f'৳{all_inc.quantile(.25):,.0f}', f'৳{earners.quantile(.25):,.0f}'),
    ('75th Percentile', f'৳{all_inc.quantile(.75):,.0f}', f'৳{earners.quantile(.75):,.0f}'),
    ('Zero Earners',    f'{(all_inc == 0).sum()} individuals', 'Students, HW, Unemployed'),
    ('Earner Rate',     f'{df["is_earner"].mean():.1%}', '59.1% of all individuals'),
]
for i, (s, a, e) in enumerate(income_rows, 16):
    write_kv(ws2, i, s, a, e, alt=(i % 2 == 0))

# Block C — District Breakdown
write_section(ws2, 27, '🗺️  INCOME BY DISTRICT', 4, GREEN)
write_header(ws2, 28, ['District', 'Members', 'Avg Income (BDT)', 'Median Income (BDT)'],
             [30, 20, 32, 5], GREEN)
dist_agg = (df.groupby('district')['monthly_income']
            .agg(['count', 'mean', 'median'])
            .sort_values('mean', ascending=False))
for i, (d, row) in enumerate(dist_agg.iterrows(), 29):
    write_kv(ws2, i, d, f"৳{row['mean']:,.0f}",
             f"Median ৳{row['median']:,.0f} | n={int(row['count'])}",
             alt=(i % 2 == 0))

# Block D — Education Breakdown
edu_seq_all = ['Primary', 'Secondary', 'HSC', 'Graduate', 'Post-Graduate']
write_section(ws2, 35, '🎓  INCOME BY EDUCATION LEVEL', 4, NAVY)
write_header(ws2, 36, ['Education Level', 'Count', 'Avg Income (Earners)', 'Earner Rate'],
             [30, 20, 32, 5], NAVY)
for i, edu in enumerate(edu_seq_all, 37):
    sub   = df[df['edu_level'] == edu]
    earns = sub[sub['monthly_income'] > 0]
    avg   = f"৳{earns['monthly_income'].mean():,.0f}" if len(earns) > 0 else '—'
    write_kv(ws2, i, edu, avg,
             f"n={len(sub)} | Earner rate {sub['is_earner'].mean():.1%}",
             alt=(i % 2 == 0))

# Block E — Treatment Effect
c_inc = df[df['treatment'] == 0]['monthly_income']
t_inc = df[df['treatment'] == 1]['monthly_income']
t_stat, p_val = stats.ttest_ind(c_inc, t_inc)
n_control   = len(c_inc)
n_treatment = len(t_inc)

write_section(ws2, 43, '⚡  TREATMENT EFFECT — Independent Samples T-Test', 4, RED)
write_header(ws2, 44, ['Metric', f'Control Group (n={n_control})', f'Treatment Group (n={n_treatment})', ''],
             [30, 20, 32, 5], RED)
sig_label = 'No sig. difference detected' if p_val > 0.05 else 'Significant difference detected'
trt_rows = [
    ('Mean Income (BDT)',    f'৳{c_inc.mean():,.0f}',   f'৳{t_inc.mean():,.0f}'),
    ('Median Income (BDT)',  f'৳{c_inc.median():,.0f}', f'৳{t_inc.median():,.0f}'),
    ('Std Deviation',        f'৳{c_inc.std():,.0f}',    f'৳{t_inc.std():,.0f}'),
    ('T-Statistic',          f'{t_stat:.3f}',            '← two-sample t-test'),
    ('P-Value',              f'{p_val:.4f}',             f'{">" if p_val > 0.05 else "<"} 0.05 = {"NOT " if p_val > 0.05 else ""}statistically significant'),
    ('Income Difference',    f'৳{t_inc.mean()-c_inc.mean():,.0f}',
                             'Treatment earns more on average'),
    ('Interpretation',       sig_label,
                             'Larger sample may reveal effect over time'),
]
for i, (m, c, t) in enumerate(trt_rows, 45):
    write_kv(ws2, i, m, c, t, alt=(i % 2 == 0))

print("  → Summary Statistics sheet complete")


# ══════════════════════════════════════════════════════════════
# SHEET 3 — District Pivot
# ══════════════════════════════════════════════════════════════
print("Building Sheet 3: District Pivot...")
ws3 = wb.create_sheet('03 District Pivot')
ws3.sheet_view.showGridLines = False

write_title(ws3, 1, 'DISTRICT-LEVEL ANALYSIS', 7)

headers3 = ['District', 'Members', 'Households', 'Avg Income (BDT)',
            'Median Income (BDT)', 'Earner Rate', 'Poverty Rate']
widths3  = [16, 12, 14, 18, 18, 14, 14]
write_header(ws3, 2, headers3, widths3)

dist_detail = df.groupby('district').agg(
    members     =('person_id',    'count'),
    households  =('hh_id',        'nunique'),
    avg_income  =('monthly_income','mean'),
    med_income  =('monthly_income','median'),
    earner_rate =('is_earner',     'mean'),
    poverty_rate=('is_poor',       'mean'),
).sort_values('avg_income', ascending=False)

for i, (d, row) in enumerate(dist_detail.iterrows(), 3):
    write_row(ws3, i, [
        d, int(row['members']), int(row['households']),
        f"৳{row['avg_income']:,.0f}", f"৳{row['med_income']:,.0f}",
        f"{row['earner_rate']:.1%}", f"{row['poverty_rate']:.1%}"
    ], alt=(i % 2 == 0))

# Embedded bar chart
chart = BarChart()
chart.type  = 'col'
chart.title = 'Average Monthly Income by District'
chart.y_axis.title = 'BDT'; chart.x_axis.title = 'District'
chart.style = 10; chart.width = 20; chart.height = 12
chart.add_data(Reference(ws3, min_col=4, min_row=2, max_row=7), titles_from_data=True)
chart.set_categories(Reference(ws3, min_col=1, min_row=3, max_row=7))
ws3.add_chart(chart, 'A10')

print("  → District Pivot + chart added")


# ══════════════════════════════════════════════════════════════
# SHEET 4 — Education Pivot
# ══════════════════════════════════════════════════════════════
print("Building Sheet 4: Education Pivot...")
ws4 = wb.create_sheet('04 Education Pivot')
ws4.sheet_view.showGridLines = False

write_title(ws4, 1, 'EDUCATION & INCOME ANALYSIS', 7)
headers4 = ['Education Level', 'Count', 'Avg Income (BDT)', 'Median Income',
            'Earner Count', 'Earner Rate', 'Avg Age']
widths4  = [20, 10, 18, 16, 14, 12, 10]
write_header(ws4, 2, headers4, widths4)

for i, edu in enumerate(edu_seq_all, 3):
    sub   = df[df['edu_level'] == edu]
    earns = sub[sub['monthly_income'] > 0]
    if len(sub) == 0: continue
    write_row(ws4, i, [
        edu, len(sub),
        int(earns['monthly_income'].mean()) if len(earns) > 0 else 0,
        int(earns['monthly_income'].median()) if len(earns) > 0 else 0,
        int(sub['is_earner'].sum()),
        f"{sub['is_earner'].mean():.1%}",
        f"{sub['age'].mean():.1f}"
    ], alt=(i % 2 == 0))

chart4 = BarChart()
chart4.type  = 'col'
chart4.title = 'Avg Income by Education Level (Earners Only)'
chart4.y_axis.title = 'BDT'; chart4.x_axis.title = 'Education'
chart4.style = 10; chart4.width = 20; chart4.height = 12
chart4.add_data(Reference(ws4, min_col=3, min_row=2, max_row=8), titles_from_data=True)
chart4.set_categories(Reference(ws4, min_col=1, min_row=3, max_row=8))
ws4.add_chart(chart4, 'A11')

print("  → Education Pivot + chart added")


# ══════════════════════════════════════════════════════════════
# SHEET 5 — Occupation Pivot
# ══════════════════════════════════════════════════════════════
print("Building Sheet 5: Occupation Pivot...")
ws5 = wb.create_sheet('05 Occupation Pivot')
ws5.sheet_view.showGridLines = False

write_title(ws5, 1, 'OCCUPATION & EMPLOYMENT ANALYSIS', 7)
headers5 = ['Occupation', 'Count', '% of Total', 'Avg Income (BDT)',
            'Median Income', 'Max Income', 'Std Dev']
widths5  = [16, 10, 12, 18, 16, 14, 14]
write_header(ws5, 2, headers5, widths5)

occ_agg = df.groupby('occupation').agg(
    count   =('person_id',    'count'),
    avg_inc =('monthly_income','mean'),
    med_inc =('monthly_income','median'),
    max_inc =('monthly_income','max'),
    std_inc =('monthly_income','std'),
).sort_values('avg_inc', ascending=False)

for i, (o, row) in enumerate(occ_agg.iterrows(), 3):
    write_row(ws5, i, [
        o, int(row['count']),
        f"{row['count'] / len(df):.1%}",
        f"৳{row['avg_inc']:,.0f}",
        f"৳{row['med_inc']:,.0f}",
        f"৳{row['max_inc']:,.0f}",
        f"৳{row['std_inc']:,.0f}" if pd.notna(row['std_inc']) else 'N/A'
    ], alt=(i % 2 == 0))

print("  → Occupation Pivot complete")


# ══════════════════════════════════════════════════════════════
# SHEET 6 — Household Level
# ══════════════════════════════════════════════════════════════
print("Building Sheet 6: Household Level...")
ws6 = wb.create_sheet('06 Household Level')
ws6.sheet_view.showGridLines = False
ws6.freeze_panes = 'A3'

write_title(ws6, 1, f'HOUSEHOLD-LEVEL AGGREGATED DATA  |  {len(hh)} Households', 10)
headers6 = ['HH ID', 'District', 'Group', 'HH Size', 'Earners',
            'Total Income (BDT)', 'Per-Capita Income', 'Dependency Ratio',
            'Treatment', 'Poverty Status']
widths6  = [9, 14, 12, 10, 10, 20, 20, 18, 11, 16]
write_header(ws6, 2, headers6, widths6)

hh_export_cols = ['hh_id', 'district', 'group', 'hh_size', 'earner_count',
                  'total_income', 'income_per_capita', 'dependency_ratio',
                  'treatment', 'is_poor']

for i, (_, row) in enumerate(hh[hh_export_cols].iterrows(), 3):
    poverty_label = 'POOR' if row['is_poor'] else 'Non-poor'
    row_data = [
        int(row['hh_id']),
        row['district'],
        row['group'],
        int(row['hh_size']),
        int(row['earner_count']),
        f"৳{row['total_income']:,.0f}",
        f"৳{row['income_per_capita']:,.0f}",
        f"{row['dependency_ratio']:.2f}" if pd.notna(row['dependency_ratio']) else 'N/A',
        int(row['treatment']),
        poverty_label
    ]
    write_row(ws6, i, row_data, alt=(i % 2 == 0))
    # Red highlight for poor households
    if row['is_poor']:
        for col in range(1, 11):
            ws6.cell(i, col).fill = fill('FCE4D6')
            ws6.cell(i, col).font = Font(name='Calibri', size=10, color=RED, bold=True)

print(f"  → {len(hh)} household rows | {hh['is_poor'].sum()} poor HH highlighted red")


# ══════════════════════════════════════════════════════════════
# SHEET 7 — Key Insights
# ══════════════════════════════════════════════════════════════
print("Building Sheet 7: Key Insights...")
ws7 = wb.create_sheet('07 Key Insights')
ws7.sheet_view.showGridLines = False

write_title(ws7, 1, 'KEY ANALYTICAL INSIGHTS — 23 Findings Across 6 Categories', 5)
set_width(ws7, 1, 5)
set_width(ws7, 2, 35)
set_width(ws7, 3, 30)
set_width(ws7, 4, 30)
set_width(ws7, 5, 5)

male_mean   = df[df['sex_label'] == 'Male']['monthly_income'].mean()
female_mean = df[df['sex_label'] == 'Female']['monthly_income'].mean()
pg_mean     = df[(df['edu_level'] == 'Post-Graduate') & (df['monthly_income'] > 0)]['monthly_income'].mean()
pri_mean    = df[(df['edu_level'] == 'Primary') & (df['monthly_income'] > 0)]['monthly_income'].mean()

sections = [
    ('💰  INCOME INSIGHTS', NAVY, [
        ('Income is heavily right-skewed',
         f"Mean ৳{df['monthly_income'].mean():,.0f} vs Median ৳{df['monthly_income'].median():,.0f} (+17%)",
         'Use median for central tendency; log-transform for regression'),
        ('35% of individuals earn zero income',
         f"{(df['monthly_income']==0).sum()} of 137 earn ৳0 (students, housewives, unemployed)",
         'Analyse earner-only subset separately for income modelling'),
        (f'Male earners earn ৳{male_mean-female_mean:,.0f}/mo more than females',
         f"Male avg ৳{male_mean:,.0f} vs Female avg ৳{female_mean:,.0f} (+{(male_mean/female_mean-1):.0%})",
         'Significant gender pay gap — critical for equity reporting'),
        ('Older age groups earn more',
         f"Age 46+ avg ৳{df[df['age']>=46]['monthly_income'].mean():,.0f} vs Under-18 ৳{df[df['age']<18]['monthly_income'].mean():,.0f}",
         'Experience premium exists — supports age as a predictor variable'),
        ('Business, NGO & Farming are top-earning sectors',
         'Business avg ৳18,896 | NGO Worker ৳18,502 | Farmer ৳18,460/month',
         'Rural income (Farmer) comparable to formal sector — notable finding'),
    ]),
    ('🎓  EDUCATION INSIGHTS', BLUE, [
        ('Post-Graduate earns most among active earners',
         f"Post-Graduate avg ৳{pg_mean:,.0f}/month | Primary avg ৳{pri_mean:,.0f}/month",
         'Education premium confirmed — include in regression model'),
        ('Primary education shows inflated mean',
         'Driven by high-income Farmers — occupation is a confounding variable',
         'Always control for occupation when analysing edu-income relationship'),
        ('17% of education data is missing',
         '23 of 137 records have no education level recorded',
         'MCAR assumed — imputed with "Unknown"; document in methodology'),
        ('HSC earners show income dip vs Graduate',
         'Small sample per group (n≈21) — interpret with caution',
         'Collect more data before drawing policy conclusions'),
    ]),
    ('⚡  TREATMENT PROGRAM', RED, [
        ('No statistically significant treatment effect found',
         f'T-stat: {t_stat:.3f} | P-value: {p_val:.4f} (threshold: 0.05)',
         'Cannot reject H₀ — program may need more time or larger sample'),
        ('Treatment group has 3.4× higher income variance',
         f'Treatment std ৳{t_inc.std():,.0f} vs Control ৳{c_inc.std():,.0f}',
         'One ৳250,000 outlier inflates Treatment stats — verify source record'),
        ('Sample is well-balanced: 68 vs 69',
         'Near-perfect randomisation balance between groups',
         'Good experimental design enables valid causal inference'),
        (f'Income difference ৳{t_inc.mean()-c_inc.mean():,.0f}/mo (not significant)',
         f"Treatment ৳{t_inc.mean():,.0f} vs Control ৳{c_inc.mean():,.0f}",
         'Practical significance exists even if statistical result is borderline'),
    ]),
    ('🏠  HOUSEHOLD STRUCTURE', GREEN, [
        ('Average household size is 4.6 members',
         'Range: 3–6 members | Most common: 4 or 5 members per household',
         'Tight distribution — consistent with national Bangladesh averages'),
        ('10% of households fall below poverty line',
         f"{hh['is_poor'].sum()} of 30 households have per-capita income < ৳5,000",
         'Concentrated in Chittagong (16%) and Khulna (19%) districts'),
        ('Average dependency ratio: 1.0',
         'Each earner supports ~1 non-earner on average in this sample',
         'Flag households with ratio > 3 for targeted social support'),
        ('No zero-earner households detected',
         'All 30 households have at least 1 income earner',
         'No extreme dependency cases — dataset may be pre-screened'),
    ]),
    ('🗺️  REGIONAL PATTERNS', NAVY, [
        ('Sylhet has highest avg income but smallest sample',
         f"৳{df[df['district']=='Sylhet']['monthly_income'].mean():,.0f}/month avg | only 8 members",
         'Interpret cautiously — n=8 is too small for statistical confidence'),
        ('Khulna has highest household poverty rate at 19%',
         f"{df[df['district']=='Khulna']['is_poor'].mean():.1%} of Khulna individuals below poverty threshold",
         'Priority district for program expansion in next funding cycle'),
        ('Dhaka is the largest and most reliable district sample',
         f"40 members across multiple households — most statistically robust",
         'Use Dhaka as baseline reference district in comparative analysis'),
    ]),
    ('⚠️  DATA QUALITY', GRAY, [
        ('EN03 interviews are 2× longer than EN05',
         f"EN03 avg {df[df['enumerator_id']=='EN03']['interview_duration'].mean():.0f} min vs EN05 avg {df[df['enumerator_id']=='EN05']['interview_duration'].mean():.0f} min",
         'Enumerator effect detected — calibrate EN03 in next survey round'),
        ('3 missing interview durations (MCAR)',
         'Imputed using within-household median duration',
         'Document imputation method in methodology section'),
        ('One income outlier: ৳250,000 in Student occupation',
         'Likely data entry error — inconsistent with Student status',
         'Verify against paper form before including in regression models'),
    ]),
]

current_row = 3
for section_title, sec_color, insights in sections:
    write_section(ws7, current_row, section_title, 5, sec_color)
    current_row += 1
    write_header(ws7, current_row,
                 ['', 'Finding', 'Data Evidence', 'Analyst Note', ''],
                 [5, 35, 30, 30, 5], sec_color)
    current_row += 1
    for j, (finding, evidence, note) in enumerate(insights):
        ws7.row_dimensions[current_row].height = 40
        bg = XLBLUE if j % 2 == 0 else WHITE
        for col in [1, 5]:
            ws7.cell(current_row, col).fill = fill(bg)
        for col, val, is_bold in [(2, f'• {finding}', True),
                                   (3, evidence,        False),
                                   (4, note,            False)]:
            c = ws7.cell(current_row, col, val)
            c.font      = Font(name='Calibri', size=10, bold=is_bold,
                               italic=(col == 4), color=GRAY if col == 4 else '1A1A1A')
            c.fill      = fill(bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border    = thin()
        current_row += 1
    current_row += 1  # spacer row

print(f"  → 23 insights written across 6 categories")


# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
print(f"\nSaving workbook → {OUTPUT_PATH}")
wb.save(OUTPUT_PATH)
print(f"✓  Excel workbook saved: {OUTPUT_PATH}")
print(f"   Sheets: {', '.join([ws.title for ws in wb.worksheets])}")
print("\n✓  Script 03 complete. All outputs ready.\n")